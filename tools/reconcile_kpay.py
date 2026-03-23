"""
KPay × DBS × POS Monthly Reconciliation Tool (v2)
===================================================
Dynamically adapts to:
  - Variable payment methods per month (columns adjust automatically)
  - Variable settle date counts (total row position adjusts)
  - Source file column layouts detected by header keywords (not hardcoded)

Output format matches the Wai Shing reference template exactly.

Usage:
    python tools/reconcile_kpay.py \
      --pos   "assets/WS_Recon/Feb2026/KYR/POS Sales_202602.xlsx" \
      --kpay  "assets/WS_Recon/Feb2026/KYR/#02_KYR_8521241461000055_202602.xlsx" \
      --dbs   "assets/WS_Recon/Feb2026/KYR/DBS_Sunstage_002378792_HKD_202602.xls" \
      --month 202602 --store KYR --merchant 146100005 \
      --output .tmp/KPay_Feb2026_KYR_Reconciliation.xlsx
"""

import argparse
import re
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta

import openpyxl
import xlrd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# MASTER CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

# Standard KPay payment methods — always shown as columns (matches reference).
# These 7 are the common methods settled via DBS.
KPAY_METHOD_ORDER = [
    'Mastercard', 'PayMe', 'Visa', '中國銀聯', '支付寶', '微信', '銀聯雲閃付',
]

# Display order for rare methods — if they appear in KPay data, they are
# appended AFTER the 7 standard columns in this order.
KPAY_RARE_METHOD_ORDER = ['轉數快', 'JCB', '八達通', 'American Express']

# POS tender name → KPay method name  (from Payment Tender Code.xlsx)
# Cash and 現金卷 map to NA — they don't go through KPay.
POS_TO_KPAY = {
    'Master':              'Mastercard',
    'Mastercard':          'Mastercard',
    'Payme':               'PayMe',
    'PayMe':               'PayMe',
    'Visa':                'Visa',
    '銀聯/BOC Pay/雲閃付': '中國銀聯',
    'UnionPay':            '中國銀聯',
    '中國銀聯':            '中國銀聯',
    '支付寶':              '支付寶',
    'Alipay':              '支付寶',
    '微信支付':            '微信',
    '微信':                '微信',
    'WeChat':              '微信',
    '銀聯雲閃付':          '銀聯雲閃付',
    '轉數快':              '轉數快',
    'FPS':                 '轉數快',
    'JCB':                 'JCB',
    '八達通':              '八達通',
    'Octopus':             '八達通',
    'AE':                  'American Express',
    'American Express':    'American Express',
}

# KPay header keywords for column auto-detection
KPAY_HEADER_MAP = {
    'txn_id':      ['KPay交易號'],
    'status':      ['交易狀態'],
    'amount':      ['支付金額'],
    'fee':         ['手續費'],
    'method':      ['支付方式'],
    'settle_date': ['清算日期'],
}

# POS header keywords for column auto-detection
POS_HEADER_MAP = {
    'store':       ['門市', 'Store', 'Shop'],
    'date':        ['日期', 'Date'],
    'Cash':        ['Cash'],
    'FPS':         ['轉數快', 'FPS'],
    'Master':      ['Master'],
    'Visa':        ['Visa'],
    'Octopus':     ['八達通', 'Octopus'],
    'PayMe':       ['Payme', 'PayMe'],
    'Alipay':      ['支付寶', 'Alipay'],
    'WeChat':      ['微信支付', 'WeChat'],
    'AE':          ['AE', 'American Express'],
    'JCB':         ['JCB'],
    'CashVoucher': ['現金卷', '現金券', 'Voucher'],
    'UnionPay':    ['銀聯/BOC Pay/雲閃付', '銀聯', 'UnionPay', 'BOC Pay'],
    'Total':       ['總計', 'Total'],
}

# ─────────────────────────────────────────────────────────────────────────────
# LAYOUT ENGINE — computes all positions from the active method list
# ─────────────────────────────────────────────────────────────────────────────

# Fixed section start columns (1-based, never change)
GRS_START = 1    # Column A  — Gross / 支付金額
FEE_START = 16   # Column P  — Fee / 手續費
NET_START = 31   # Column AE — Net Amount

# Fixed row positions
ROW_SECTION_HDR = 3
ROW_COL_HDR     = 4
ROW_DATA_START  = 5
ROW_POS         = 38   # Always row 38
ROW_DIFF        = 39   # Always row 39


def compute_layout(methods):
    """
    Compute all dynamic column positions from the active payment method list.

    Section layout (each section = 1 label + N methods + 1 total = N+2 cols):
      GRS: A(1)  … always fixed
      FEE: P(16) … always fixed
      NET: AE(31) … always fixed
      DBS: 8 cols after NET section end  (3 cols: Receipt, Amount, Variance)
      RATE: 6 cols after DBS start       (N+2 cols, same as other sections)
    """
    n = len(methods)
    # NET section ends at column: NET_START + n + 1
    # DBS starts 8 columns after that
    dbs_start = NET_START + n + 1 + 8          # = n + 40
    rate_start = dbs_start + 6                  # DBS(3 cols) + 3 gap cols
    return {
        'methods':     methods,
        'num_methods': n,
        'GRS_START':   GRS_START,
        'FEE_START':   FEE_START,
        'NET_START':   NET_START,
        'DBS_START':   dbs_start,
        'RATE_START':  rate_start,
        'total_cols':  rate_start + n + 1,
    }


def compute_rows(num_settle_dates):
    """
    Compute row positions from the number of KPay settlement dates.

    The template reserves at least 30 data rows (rows 5-34) so the total
    row is at 35 minimum.  Months with 31+ settle dates push it to 36+.
    """
    last_data_row = ROW_DATA_START + num_settle_dates - 1
    # Pad to at least 30 data rows (standard month)
    total_row = max(last_data_row + 1, ROW_DATA_START + 30)
    return {
        'last_data_row': last_data_row,
        'ROW_TOTAL':     total_row,
        'ROW_POS':       ROW_POS,
        'ROW_DIFF':      ROW_DIFF,
    }


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def fmt_date(d):
    """date/datetime → 'dd/mm/yyyy'."""
    if isinstance(d, datetime):
        return d.strftime('%d/%m/%Y')
    if isinstance(d, date):
        return d.strftime('%d/%m/%Y')
    return str(d)


def fmt_dbs_date(d):
    """date → 'dd-Mon-yyyy' for the DBS receipt column."""
    if hasattr(d, 'strftime'):
        return d.strftime('%d-%b-%Y')
    return str(d)


def parse_date_str(s):
    """Parse various date formats → date object, or None."""
    if isinstance(s, datetime):
        return s.date()
    if isinstance(s, date):
        return s
    s = str(s).strip()
    for fmt in ('%d/%m/%Y', '%d-%b-%Y', '%Y-%m-%d', '%d-%B-%Y',
                '%Y/%m/%d', '%d/%m/%y'):
        try:
            return datetime.strptime(s, fmt).date()
        except (ValueError, TypeError):
            pass
    return None


def col_ltr(n):
    """1-based column number → Excel letter(s)."""
    return get_column_letter(n)


# HK general public holidays for T+2 business-day calculation.
# Source: https://www.gov.hk/en/about/abouthk/holiday/
# Only weekday holidays matter (weekends already skipped by weekday check),
# but we include all dates for completeness and clarity.
# UPDATE THIS LIST each year when the government gazettes the next year's holidays.
HK_HOLIDAYS = {
    # ── 2025 (verified from gov.hk) ──────────────────────────────────────
    date(2025, 1, 1),                                       # New Year's Day (Wed)
    date(2025, 1, 29), date(2025, 1, 30), date(2025, 1, 31),  # Lunar New Year (Wed-Fri)
    date(2025, 4, 4),                                       # Ching Ming Festival (Fri)
    date(2025, 4, 18), date(2025, 4, 19), date(2025, 4, 21),  # Good Friday (Fri), Sat, Easter Mon
    date(2025, 5, 1),                                       # Labour Day (Thu)
    date(2025, 5, 5),                                       # Buddha's Birthday (Mon)
    date(2025, 5, 31),                                      # Tuen Ng Festival (Sat — no weekday replacement)
    date(2025, 7, 1),                                       # HKSAR Establishment Day (Tue)
    date(2025, 10, 1),                                      # National Day (Wed)
    date(2025, 10, 7),                                      # Day after Mid-Autumn (Tue)
    date(2025, 10, 29),                                     # Chung Yeung Festival (Wed)
    date(2025, 12, 25), date(2025, 12, 26),                 # Christmas (Thu-Fri)
    # ── 2026 (verified from gov.hk) ──────────────────────────────────────
    date(2026, 1, 1),                                       # New Year's Day (Thu)
    date(2026, 2, 17), date(2026, 2, 18), date(2026, 2, 19),  # Lunar New Year (Tue-Thu)
    date(2026, 4, 3), date(2026, 4, 4),                     # Good Friday (Fri), day after (Sat)
    date(2026, 4, 6), date(2026, 4, 7),                     # Ching Ming replacement (Mon), Easter Mon replacement (Tue)
    date(2026, 5, 1),                                       # Labour Day (Fri)
    date(2026, 5, 25),                                      # Buddha's Birthday replacement (Mon; 24th is Sun)
    date(2026, 6, 19),                                      # Tuen Ng Festival (Fri)
    date(2026, 7, 1),                                       # HKSAR Establishment Day (Wed)
    date(2026, 9, 26),                                      # Day after Mid-Autumn (Sat)
    date(2026, 10, 1),                                      # National Day (Thu)
    date(2026, 10, 19),                                     # Chung Yeung replacement (Mon; 18th is Sun)
    date(2026, 12, 25), date(2026, 12, 26),                 # Christmas (Fri), day after (Sat)
    # ── 2027 (provisional — verify when gazetted) ────────────────────────
    date(2027, 1, 1),                                       # New Year's Day (Fri)
    date(2027, 2, 6), date(2027, 2, 8), date(2027, 2, 9),  # Lunar New Year (Sat→replacement Mon/Tue)
    date(2027, 3, 26), date(2027, 3, 27), date(2027, 3, 29),  # Good Friday (Fri), Sat, Easter Mon
    date(2027, 4, 5),                                       # Ching Ming Festival (Mon)
    date(2027, 5, 1),                                       # Labour Day (Sat — no weekday replacement TBC)
    date(2027, 5, 13),                                      # Buddha's Birthday (Thu)
    date(2027, 6, 9),                                       # Tuen Ng Festival (Wed)
    date(2027, 7, 1),                                       # HKSAR Establishment Day (Thu)
    date(2027, 9, 16),                                      # Day after Mid-Autumn (Thu)
    date(2027, 10, 1),                                      # National Day (Fri)
    date(2027, 10, 8),                                      # Chung Yeung Festival (Fri)
    date(2027, 12, 25), date(2027, 12, 27),                 # Christmas (Sat→Mon replacement)
}

# Range of years we have holiday data for
_HK_HOLIDAY_YEARS = {d.year for d in HK_HOLIDAYS}


def next_biz_day(d, n=2):
    """Return d + n business days, skipping weekends and HK public holidays."""
    if d.year not in _HK_HOLIDAY_YEARS:
        print(f'  WARNING: No HK holiday data for {d.year} — T+2 may be inaccurate. '
              f'Update HK_HOLIDAYS in reconcile_kpay.py.')
    count, cur = 0, d
    while count < n:
        cur += timedelta(days=1)
        if cur.weekday() < 5 and cur not in HK_HOLIDAYS:
            count += 1
    return cur


# ─────────────────────────────────────────────────────────────────────────────
# COLUMN DETECTION
# ─────────────────────────────────────────────────────────────────────────────

def _find_header_row(rows, keywords, max_rows=20):
    """
    Scan rows to find the header row containing the given keyword sets.
    Returns (row_index, {field: 0-based col_index}) or (None, {}).
    """
    for ri, row in enumerate(rows[:max_rows]):
        matches = {}
        row_strs = [str(v).strip() if v else '' for v in row]
        for field, kws in keywords.items():
            for ci, cell_val in enumerate(row_strs):
                if cell_val and any(kw in cell_val for kw in kws):
                    matches[field] = ci
                    break
        if len(matches) >= 3:
            return ri, matches
    return None, {}


def detect_kpay_columns(ws_or_rows):
    """Detect KPay column positions by scanning for header keywords."""
    if hasattr(ws_or_rows, 'iter_rows'):
        rows = list(ws_or_rows.iter_rows(min_row=1, max_row=20, values_only=True))
    else:
        rows = ws_or_rows[:20]
    return _find_header_row(rows, KPAY_HEADER_MAP)


def detect_pos_columns(ws):
    """Detect POS column positions by scanning header rows for keywords."""
    rows = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
    result = {}
    for row in rows:
        row_strs = [str(v).strip() if v else '' for v in row]
        for field, kws in POS_HEADER_MAP.items():
            if field in result:
                continue
            for ci, cell_val in enumerate(row_strs):
                if cell_val and any(kw.lower() in cell_val.lower() for kw in kws):
                    result[field] = ci
                    break
    return result


# ─────────────────────────────────────────────────────────────────────────────
# DATA READERS
# ─────────────────────────────────────────────────────────────────────────────

def read_kpay(kpay_file):
    """
    Read KPay settlement transactions with dynamic column detection.

    Returns:
        by_settle    : {date_str: {method: {'gross': float, 'fee': float}}}
        settle_dates : sorted list of date_str  (dd/mm/yyyy)
        methods_found: set of payment method names actually present in data
    """
    methods_found = set()
    by_settle = defaultdict(lambda: defaultdict(lambda: {'gross': 0.0, 'fee': 0.0}))

    if str(kpay_file).lower().endswith('.xls'):
        xwb = xlrd.open_workbook(kpay_file)
        try:
            xws = xwb.sheet_by_name('交易結算')
        except xlrd.XLRDError:
            xws = xwb.sheet_by_index(0)

        rows_data = [xws.row_values(i) for i in range(xws.nrows)]
        hdr_ri, col_map = _find_header_row(rows_data, KPAY_HEADER_MAP)

        if hdr_ri is None:
            print('  WARNING: Could not detect KPay headers, using fallback')
            hdr_ri = 8
            col_map = {'txn_id': 0, 'status': 6, 'amount': 8,
                       'fee': 13, 'method': 18, 'settle_date': 24}
        else:
            print(f'  KPay header detected at row {hdr_ri + 1}')

        for row in rows_data[hdr_ri + 1:]:
            if not row[col_map['txn_id']]:
                continue
            if str(row[col_map['status']]).strip() != '處理成功':
                continue
            method = str(row[col_map['method']]).strip()
            gross  = float(row[col_map['amount']] or 0)
            fee    = float(row[col_map['fee']]    or 0)
            raw_sd = row[col_map['settle_date']]
            if isinstance(raw_sd, float) and raw_sd > 0:
                sd = xlrd.xldate_as_datetime(raw_sd, xwb.datemode)
                ds = fmt_date(sd)
            else:
                d = parse_date_str(raw_sd)
                ds = fmt_date(d) if d else ''
            if not ds:
                continue
            methods_found.add(method)
            by_settle[ds][method]['gross'] += gross
            by_settle[ds][method]['fee']   += fee

    else:  # .xlsx
        wb = openpyxl.load_workbook(kpay_file)
        ws = None
        for name in ['交易結算', 'KPay交易結算']:
            if name in wb.sheetnames:
                ws = wb[name]
                break
        if ws is None:
            ws = wb.active

        hdr_ri, col_map = detect_kpay_columns(ws)
        if hdr_ri is None:
            print('  WARNING: Could not detect KPay headers, using fallback')
            hdr_ri = 8
            col_map = {'txn_id': 0, 'status': 6, 'amount': 8,
                       'fee': 13, 'method': 18, 'settle_date': 24}
        else:
            print(f'  KPay header detected at row {hdr_ri + 1}')

        for row in ws.iter_rows(min_row=hdr_ri + 2, values_only=True):
            if row[col_map['txn_id']] is None:
                continue
            status = str(row[col_map['status']]).strip() if row[col_map['status']] else ''
            if status != '處理成功':
                continue
            method = str(row[col_map['method']]).strip() if row[col_map['method']] else ''
            amount = row[col_map['amount']] or 0
            fee    = row[col_map['fee']]    or 0
            sd_raw = row[col_map['settle_date']]
            if not sd_raw or not method:
                continue
            d = sd_raw if isinstance(sd_raw, (date, datetime)) else parse_date_str(sd_raw)
            ds = fmt_date(d) if d else ''
            if not ds:
                continue
            methods_found.add(method)
            by_settle[ds][method]['gross'] += amount
            by_settle[ds][method]['fee']   += fee
        wb.close()

    settle_dates = sorted(by_settle.keys(),
                          key=lambda s: parse_date_str(s) or date.min)
    return dict(by_settle), settle_dates, methods_found


def read_pos(pos_file, store_code):
    """
    Read POS daily sales with dynamic column detection.

    Returns:
        daily   : {date_str: {kpay_method_name: amount}}
        non_kpay: {pos_tender: monthly_total}
    """
    wb = openpyxl.load_workbook(pos_file)
    for name in ['Sheet (2)', 'Sheet']:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    else:
        ws = wb.active

    PC = detect_pos_columns(ws)
    print(f'  POS columns detected: { {k: v for k, v in PC.items() if k in ["store","date","Master","Visa","PayMe","Alipay","WeChat","UnionPay"]} }')

    # Auto-detect the POS store code: scan all unique store values and match
    # against the requested code.  Handles MI→MIKI, MO→MOKO, etc.
    pos_store = store_code
    if 'store' in PC:
        all_stores = set()
        for row in ws.iter_rows(min_row=3, values_only=True):
            v = str(row[PC['store']]).strip() if row[PC['store']] else ''
            if v and 'Total' not in v and v != '總計':
                all_stores.add(v)
        if store_code not in all_stores:
            # Try prefix match (MI→MIKI, MO→MOKO)
            candidates = [s for s in all_stores if s.startswith(store_code)]
            if len(candidates) == 1:
                pos_store = candidates[0]
                print(f'  POS store code mapped: {store_code} → {pos_store}')
            elif candidates:
                pos_store = min(candidates, key=len)
                print(f'  POS store code mapped: {store_code} → {pos_store} (from {candidates})')
            else:
                print(f'  WARNING: Store code "{store_code}" not found in POS. '
                      f'Available: {sorted(all_stores)}')

    # All POS fields that map to a KPay method (via POS_TO_KPAY)
    kpay_fields  = ['Master', 'PayMe', 'Visa', 'UnionPay', 'Alipay', 'WeChat',
                    'FPS', 'JCB', 'Octopus', 'AE']
    # POS fields that do NOT go through KPay
    nk_fields    = ['Cash', 'CashVoucher']
    daily = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        store_val = str(row[PC['store']]).strip() if PC.get('store') is not None and row[PC['store']] else ''
        if store_val != pos_store:
            continue
        d = row[PC['date']] if 'date' in PC else None
        if d is None:
            continue
        ds = fmt_date(d)

        day = {}
        for field in kpay_fields:
            if field in PC and PC[field] < len(row):
                kpay_name = POS_TO_KPAY.get(field, field)
                day[kpay_name] = row[PC[field]] or 0
        daily[ds] = day

    non_kpay = {}
    for field in nk_fields:
        if field in PC:
            non_kpay[field] = sum(
                (row_data.get(field, 0) or 0)
                for row_data in daily.values()
            )

    wb.close()
    return daily, non_kpay


def read_dbs(dbs_file, merchant_id, year, month):
    """
    Read DBS bank statement (.xls or .xlsx) filtered to exact month & merchant.

    Returns:
        by_date: {date_str: total_credit}
        batches: list of {date, credit, batch_no}
    """
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)

    # ── .xls path (xlrd) ─────────────────────────────────────────────────
    if str(dbs_file).lower().endswith('.xls'):
        xls = xlrd.open_workbook(dbs_file)
        ws  = xls.sheet_by_index(0)

        # Find the first data row by scanning for a date value in column A
        data_start = None
        for r in range(ws.nrows):
            cell = ws.cell(r, 0)
            if cell.ctype == xlrd.XL_CELL_DATE:
                data_start = r
                break
            if cell.ctype == xlrd.XL_CELL_TEXT and parse_date_str(cell.value):
                data_start = r
                break
        if data_start is None:
            data_start = 6
            print(f'  WARNING: Could not detect DBS data start, using row {data_start + 1}')
        else:
            print(f'  DBS data starts at row {data_start + 1}')

        by_date = defaultdict(float)
        batches = []

        for r in range(data_start, ws.nrows):
            row = [ws.cell_value(r, c) for c in range(min(ws.ncols, 10))]

            # Combine description columns for KPAY / merchant-ID search
            desc_parts = [str(row[c]) for c in range(2, min(5, len(row))) if row[c]]
            desc = ' '.join(desc_parts).upper()
            if 'KPAY' not in desc or merchant_id not in desc:
                continue

            # Parse date
            date_raw = row[0]
            if isinstance(date_raw, float) and date_raw > 0:
                try:
                    tup = xlrd.xldate_as_tuple(date_raw, xls.datemode)
                    d = date(*tup[:3])
                except Exception:
                    continue
            else:
                d = parse_date_str(date_raw)
            if d is None or not (start_date <= d <= end_date):
                continue

            # Find credit amount (positive number in cols 4-6)
            credit = 0
            for ci in [5, 4, 6]:
                if ci < len(row) and isinstance(row[ci], (int, float)) and row[ci] > 0:
                    credit = row[ci]
                    break
            if credit == 0:
                continue

            ds = fmt_date(d)
            by_date[ds] += credit
            batch_match = re.search(rf'{re.escape(merchant_id)}K(\d+)', desc)
            batch_no = batch_match.group(1) if batch_match else ''
            batches.append({'date': ds, 'credit': credit, 'batch_no': batch_no})

        return dict(by_date), batches

    # ── .xlsx path (openpyxl) ─────────────────────────────────────────────
    wb = openpyxl.load_workbook(dbs_file)
    ws = wb.active
    by_date = defaultdict(float)
    batches = []

    for row in ws.iter_rows(min_row=7, values_only=True):
        desc_parts = [str(row[c]) for c in range(2, min(5, len(row))) if row[c]]
        desc = ' '.join(desc_parts).upper()
        if 'KPAY' not in desc or merchant_id not in desc:
            continue
        d = row[0] if isinstance(row[0], (date, datetime)) else parse_date_str(row[0])
        if isinstance(d, datetime):
            d = d.date()
        if d is None or not (start_date <= d <= end_date):
            continue
        credit = 0
        for ci in [5, 4, 6]:
            if ci < len(row) and isinstance(row[ci], (int, float)) and row[ci] > 0:
                credit = row[ci]
                break
        if credit == 0:
            continue
        ds = fmt_date(d)
        by_date[ds] += credit
        batch_match = re.search(rf'{re.escape(merchant_id)}K(\d+)', desc)
        batch_no = batch_match.group(1) if batch_match else ''
        batches.append({'date': ds, 'credit': credit, 'batch_no': batch_no})

    wb.close()
    return dict(by_date), batches


# ─────────────────────────────────────────────────────────────────────────────
# DETERMINE ACTIVE PAYMENT METHODS
# ─────────────────────────────────────────────────────────────────────────────

def determine_methods(methods_found):
    """
    Split payment methods into standard (main grid) and rare (separate section).

    Returns:
        methods : list of 7 standard methods (always the same, matches template)
        rare    : list of rare methods actually found in data (e.g. FPS, JCB)
    """
    methods = list(KPAY_METHOD_ORDER)  # always the 7 standard ones

    # Identify rare methods that actually appear in the data
    extras = [m for m in methods_found if m not in methods]
    rare_rank = {m: i for i, m in enumerate(KPAY_RARE_METHOD_ORDER)}
    extras.sort(key=lambda m: (rare_rank.get(m, 999), m))

    for m in extras:
        print(f'  NOTE: Rare payment method found: {m} (will show separately)')

    return methods, extras


# ─────────────────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _hdr(cell, text=None, bg='1F4E79', fg='FFFFFF', bold=True, size=9):
    if text is not None:
        cell.value = text
    cell.font      = Font(name='Arial', bold=bold, color=fg, size=size)
    cell.fill      = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center',
                               wrap_text=True)


def _num(cell, val=None, fmt='#,##0.00', color='000000', bold=False):
    if val is not None:
        cell.value = val
    cell.number_format = fmt
    cell.font      = Font(name='Arial', color=color, size=9, bold=bold)
    cell.alignment = Alignment(horizontal='right', vertical='center')


def _total_style(cell):
    cell.font  = Font(name='Arial', bold=True, size=9)
    cell.fill  = PatternFill('solid', fgColor='FFF2CC')
    cell.alignment = Alignment(horizontal='right', vertical='center')


def _lbl(cell, text, bold=False):
    cell.value = text
    cell.font  = Font(name='Arial', bold=bold, size=9)
    cell.alignment = Alignment(horizontal='center', vertical='center')


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL SHEET BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def build_sheet(wb, pos_daily, kpay_by_settle, settle_dates, dbs_by_date,
                store_code, year, month, methods=None, rare_methods=None):
    """
    Build the 'KPay-By Date_Reconciliation' sheet with fully dynamic layout.

    *methods*      — list of standard methods for the main grid (default: 7).
    *rare_methods* — list of rare methods to show in a separate section below
                     the main grid (e.g. FPS, JCB).
    Returns (layout_dict, rows_dict).
    """
    if methods is None:
        methods = list(KPAY_METHOD_ORDER)
    if rare_methods is None:
        rare_methods = []

    layout = compute_layout(methods)
    rows   = compute_rows(len(settle_dates))

    n    = layout['num_methods']
    GRS  = layout['GRS_START']
    FEE  = layout['FEE_START']
    NET  = layout['NET_START']
    DBS  = layout['DBS_START']
    RATE = layout['RATE_START']
    ROW_TOTAL = rows['ROW_TOTAL']
    last_data = rows['last_data_row']

    ws = wb.create_sheet('KPay-By Date_Reconciliation')

    # ── Row 3: section headers (merged) ───────────────────────────────────
    for start, label in [(GRS,  '加總 - 支付金額'),
                         (FEE,  '加總 - 手續費'),
                         (NET,  '加總 - Net Amount'),
                         (RATE, 'Checking on Credit Card Charge Rate')]:
        _hdr(ws.cell(ROW_SECTION_HDR, start), label)
        ws.merge_cells(start_row=ROW_SECTION_HDR, start_column=start,
                       end_row=ROW_SECTION_HDR,   end_column=start + n + 1)

    _hdr(ws.cell(ROW_SECTION_HDR, DBS), 'DBS Receipt', bg='833C00')
    ws.merge_cells(start_row=ROW_SECTION_HDR, start_column=DBS,
                   end_row=ROW_SECTION_HDR,   end_column=DBS + 2)
    ws.row_dimensions[ROW_SECTION_HDR].height = 25

    # ── Row 4: column sub-headers ─────────────────────────────────────────
    for sec in [GRS, FEE, NET, RATE]:
        _hdr(ws.cell(ROW_COL_HDR, sec), '列標籤', bg='2F75B6')
        for i, m in enumerate(methods):
            _hdr(ws.cell(ROW_COL_HDR, sec + 1 + i), m, bg='2F75B6')
        _hdr(ws.cell(ROW_COL_HDR, sec + n + 1), '總計', bg='2F75B6')

    for ci, lbl_text in enumerate(['DBS Receipt', 'DBS Receipt2', 'Txn Charge']):
        _hdr(ws.cell(ROW_COL_HDR, DBS + ci), lbl_text, bg='833C00')
    ws.row_dimensions[ROW_COL_HDR].height = 30

    # ── Rows 5+: daily KPay data ──────────────────────────────────────────
    for ri, ds in enumerate(settle_dates):
        r = ROW_DATA_START + ri
        if r >= ROW_POS:
            break  # safety

        day = kpay_by_settle.get(ds, {})
        gross_vals = [day.get(m, {}).get('gross', 0) or None for m in methods]
        fee_vals   = [day.get(m, {}).get('fee',   0) or None for m in methods]

        # — Gross
        _lbl(ws.cell(r, GRS), ds)
        for i, v in enumerate(gross_vals):
            _num(ws.cell(r, GRS + 1 + i), v)
        tot = ws.cell(r, GRS + n + 1)
        tot.value = f'=SUM({col_ltr(GRS+1)}{r}:{col_ltr(GRS+n)}{r})'
        _num(tot)

        # — Fee
        ws.cell(r, FEE).value = f'={col_ltr(GRS)}{r}'
        ws.cell(r, FEE).font  = Font(name='Arial', size=9)
        ws.cell(r, FEE).alignment = Alignment(horizontal='center')
        for i, v in enumerate(fee_vals):
            _num(ws.cell(r, FEE + 1 + i), v)
        tot = ws.cell(r, FEE + n + 1)
        tot.value = f'=SUM({col_ltr(FEE+1)}{r}:{col_ltr(FEE+n)}{r})'
        _num(tot)

        # — Net  (Gross − Fee per method)
        ws.cell(r, NET).value = f'={col_ltr(GRS)}{r}'
        ws.cell(r, NET).font  = Font(name='Arial', size=9)
        ws.cell(r, NET).alignment = Alignment(horizontal='center')
        for i in range(n):
            g = col_ltr(GRS + 1 + i)
            f = col_ltr(FEE + 1 + i)
            c = ws.cell(r, NET + 1 + i)
            c.value = f'={g}{r}-{f}{r}'
            _num(c)
        tot = ws.cell(r, NET + n + 1)
        tot.value = f'=SUM({col_ltr(NET+1)}{r}:{col_ltr(NET+n)}{r})'
        _num(tot)

        # — Rate  (Fee ÷ Gross per method, as %)
        ws.cell(r, RATE).value = f'={col_ltr(GRS)}{r}'
        ws.cell(r, RATE).font  = Font(name='Arial', size=9)
        ws.cell(r, RATE).alignment = Alignment(horizontal='center')
        for i in range(n):
            g = col_ltr(GRS + 1 + i)
            f = col_ltr(FEE + 1 + i)
            c = ws.cell(r, RATE + 1 + i)
            c.value = f'=IFERROR({f}{r}/{g}{r},"-")'
            c.number_format = '0.00%'
            c.font = Font(name='Arial', size=9)
            c.alignment = Alignment(horizontal='right')
        tot_r = ws.cell(r, RATE + n + 1)
        tot_r.value = (f'=IFERROR({col_ltr(FEE+n+1)}{r}'
                       f'/{col_ltr(GRS+n+1)}{r},"-")')
        tot_r.number_format = '0.00%'
        tot_r.font = Font(name='Arial', size=9)
        tot_r.alignment = Alignment(horizontal='right')

    # ── T+2 map (KPay settle date → expected DBS receipt date) ────────────
    t2_map = defaultdict(list)
    for ri, ds in enumerate(settle_dates):
        kpay_r = ROW_DATA_START + ri
        if kpay_r >= ROW_POS:
            break
        sd = parse_date_str(ds)
        if sd:
            t2_map[fmt_date(next_biz_day(sd, 2))].append(kpay_r)

    # ── DBS receipt data ──────────────────────────────────────────────────
    net_tot_col = col_ltr(NET + n + 1)
    av_col      = col_ltr(DBS + 1)

    last_dbs_row        = ROW_DATA_START - 1
    last_match_dbs_row  = None

    sorted_dbs = sorted(dbs_by_date.items(),
                        key=lambda x: parse_date_str(x[0]) or date.min)
    for dbs_ri, (d_str, credit) in enumerate(sorted_dbs):
        r = ROW_DATA_START + dbs_ri
        if r >= ROW_POS:
            break
        last_dbs_row = r
        d_obj = parse_date_str(d_str)

        ws.cell(r, DBS).value     = fmt_dbs_date(d_obj) if d_obj else d_str
        ws.cell(r, DBS).font      = Font(name='Arial', size=9)
        ws.cell(r, DBS).alignment = Alignment(horizontal='center')

        ws.cell(r, DBS + 1).value = credit
        _num(ws.cell(r, DBS + 1))

        kpay_rows = t2_map.get(d_str, [])
        if kpay_rows:
            net_refs = '+'.join(f'{net_tot_col}{kr}' for kr in kpay_rows)
            ws.cell(r, DBS + 2).value = f'={net_refs}-{av_col}{r}'
            _num(ws.cell(r, DBS + 2))
            last_match_dbs_row = r

    # ── Total row (總計) ──────────────────────────────────────────────────
    for sec in [GRS, FEE, NET]:
        c35 = ws.cell(ROW_TOTAL, sec)
        c35.value = '總計'
        _total_style(c35)
        c35.alignment = Alignment(horizontal='right', vertical='center')
        for i in range(n + 1):
            col = col_ltr(sec + 1 + i)
            c = ws.cell(ROW_TOTAL, sec + 1 + i)
            c.value = f'=SUM({col}{ROW_DATA_START}:{col}{last_data})'
            _total_style(c)
            _num(c)

    # Rate total row
    ws.cell(ROW_TOTAL, RATE).value = '總計'
    _total_style(ws.cell(ROW_TOTAL, RATE))
    ws.cell(ROW_TOTAL, RATE).alignment = Alignment(horizontal='right',
                                                    vertical='center')
    for i in range(n):
        g = col_ltr(GRS + 1 + i)
        f = col_ltr(FEE + 1 + i)
        c = ws.cell(ROW_TOTAL, RATE + 1 + i)
        c.value = f'=IFERROR({f}{ROW_TOTAL}/{g}{ROW_TOTAL},"-")'
        c.number_format = '0.00%'
        _total_style(c)
    tr = ws.cell(ROW_TOTAL, RATE + n + 1)
    tr.value = (f'=IFERROR({col_ltr(FEE+n+1)}{ROW_TOTAL}'
                f'/{col_ltr(GRS+n+1)}{ROW_TOTAL},"-")')
    tr.number_format = '0.00%'
    _total_style(tr)

    # DBS total row
    ws.cell(ROW_TOTAL, DBS).value = '總計'
    _total_style(ws.cell(ROW_TOTAL, DBS))
    ws.cell(ROW_TOTAL, DBS).alignment = Alignment(horizontal='right')
    dbs_amt = col_ltr(DBS + 1)
    ws.cell(ROW_TOTAL, DBS + 1).value = (
        f'=SUM({dbs_amt}{ROW_DATA_START}:{dbs_amt}{last_dbs_row})')
    _total_style(ws.cell(ROW_TOTAL, DBS + 1))
    _num(ws.cell(ROW_TOTAL, DBS + 1))
    aw = col_ltr(DBS + 2)
    ws.cell(ROW_TOTAL, DBS + 2).value = (
        f'=SUM({aw}{ROW_DATA_START}:{aw}{last_dbs_row})')
    _total_style(ws.cell(ROW_TOTAL, DBS + 2))
    _num(ws.cell(ROW_TOTAL, DBS + 2))

    # ── Row 38: POS monthly totals ────────────────────────────────────────
    pos_totals = {m: sum((dd.get(m, 0) or 0) for dd in pos_daily.values())
                  for m in methods}

    ws.cell(ROW_POS, GRS).value = f'POS ({store_code})'
    _hdr(ws.cell(ROW_POS, GRS), bg='833C00')
    for i, m in enumerate(methods):
        c = ws.cell(ROW_POS, GRS + 1 + i)
        c.value = pos_totals[m]
        _num(c, color='0000FF')
    ws.cell(ROW_POS, GRS + n + 1).value = (
        f'=SUM({col_ltr(GRS+1)}{ROW_POS}:{col_ltr(GRS+n)}{ROW_POS})')
    _num(ws.cell(ROW_POS, GRS + n + 1))

    # Net total on row 38
    ws.cell(ROW_POS, NET + n + 1).value = (
        f'=SUM({net_tot_col}{ROW_DATA_START}:{net_tot_col}{last_data})')
    _num(ws.cell(ROW_POS, NET + n + 1))

    # DBS on row 38
    ws.cell(ROW_POS, DBS).value = 'Per DBS'
    _hdr(ws.cell(ROW_POS, DBS), bg='833C00')
    last_av = last_match_dbs_row if last_match_dbs_row else last_dbs_row
    ws.cell(ROW_POS, DBS + 1).value = (
        f'=SUM({av_col}{ROW_DATA_START}:{av_col}{last_av})')
    _num(ws.cell(ROW_POS, DBS + 1))
    ws.cell(ROW_POS, DBS + 2).value = (
        f'=+{net_tot_col}{ROW_POS}-{av_col}{ROW_POS}')
    _num(ws.cell(ROW_POS, DBS + 2))

    # ── Row 39: Difference  (KPay total − POS) ───────────────────────────
    ws.cell(ROW_DIFF, GRS).value = 'Difference'
    _hdr(ws.cell(ROW_DIFF, GRS), bg='C00000')
    for i in range(n + 1):
        col = col_ltr(GRS + 1 + i)
        c = ws.cell(ROW_DIFF, GRS + 1 + i)
        c.value = f'={col}{ROW_TOTAL}-{col}{ROW_POS}'
        c.number_format = '#,##0.00;(#,##0.00);"-"'
        c.font = Font(name='Arial', bold=True, size=9)

    # DBS checking row
    ws.cell(ROW_DIFF, DBS).value = 'Checking'
    _hdr(ws.cell(ROW_DIFF, DBS), bg='833C00')
    ws.cell(ROW_DIFF, DBS + 1).value = f'={av_col}{ROW_POS}-{av_col}{ROW_TOTAL}'
    _num(ws.cell(ROW_DIFF, DBS + 1))

    # ── Rare methods summary (below main grid) ───────────────────────────
    if rare_methods:
        rare_totals = {}  # {method: {'gross': x, 'fee': y}}
        for m in rare_methods:
            g = sum(kpay_by_settle.get(ds, {}).get(m, {}).get('gross', 0)
                    for ds in settle_dates)
            f = sum(kpay_by_settle.get(ds, {}).get(m, {}).get('fee', 0)
                    for ds in settle_dates)
            rare_totals[m] = {'gross': g, 'fee': f}

        r_start = ROW_DIFF + 2  # row 41
        _hdr(ws.cell(r_start, GRS), 'Other Payment Methods (not in main grid)',
             bg='7030A0')
        ws.merge_cells(start_row=r_start, start_column=GRS,
                       end_row=r_start, end_column=GRS + 5)

        _hdr(ws.cell(r_start + 1, GRS), 'Method',     bg='7030A0')
        _hdr(ws.cell(r_start + 1, GRS + 1), 'Gross',  bg='7030A0')
        _hdr(ws.cell(r_start + 1, GRS + 2), 'Fee',    bg='7030A0')
        _hdr(ws.cell(r_start + 1, GRS + 3), 'Net',    bg='7030A0')
        _hdr(ws.cell(r_start + 1, GRS + 4), 'Txn Days', bg='7030A0')

        for ri, m in enumerate(rare_methods):
            r = r_start + 2 + ri
            t = rare_totals[m]
            txn_days = sum(1 for ds in settle_dates
                          if kpay_by_settle.get(ds, {}).get(m, {}).get('gross', 0) > 0)
            _lbl(ws.cell(r, GRS), m, bold=True)
            _num(ws.cell(r, GRS + 1), t['gross'])
            _num(ws.cell(r, GRS + 2), t['fee'])
            _num(ws.cell(r, GRS + 3), t['gross'] - t['fee'])
            ws.cell(r, GRS + 4).value = txn_days
            ws.cell(r, GRS + 4).font = Font(name='Arial', size=9)
            ws.cell(r, GRS + 4).alignment = Alignment(horizontal='right')

    return layout, rows


# ─────────────────────────────────────────────────────────────────────────────
# VALIDATION
# ─────────────────────────────────────────────────────────────────────────────

def validate_output(output_file, year, month):
    """
    Validate the generated reconciliation file.
    Returns (all_passed, [(check_name, ok, detail), ...]).
    """
    results = []
    wb = openpyxl.load_workbook(output_file)

    # 1 — Sheet name
    sheet = 'KPay-By Date_Reconciliation'
    ok = sheet in wb.sheetnames
    results.append(('Sheet name', ok, f'Found: {wb.sheetnames}' if not ok else ''))
    if not ok:
        wb.close()
        return False, results
    ws = wb[sheet]

    # 2 — Section headers in row 3
    r3 = [str(ws.cell(3, c).value) for c in range(1, ws.max_column + 1)
          if ws.cell(3, c).value]
    ok = (any('支付金額' in v for v in r3) and
          any('手續費'   in v for v in r3) and
          any('Net Amount' in v for v in r3))
    results.append(('Section headers (Row 3)', ok, ''))

    # 3 — Payment-method headers in row 4
    r4 = [str(ws.cell(4, c).value) for c in range(1, ws.max_column + 1)
          if ws.cell(4, c).value]
    ok = any('Mastercard' in v for v in r4)
    results.append(('Payment method headers (Row 4)', ok, ''))

    # 4 — Find 總計 row
    total_row = None
    for r in range(ROW_DATA_START, ROW_POS):
        v = ws.cell(r, 1).value
        if v and '總計' in str(v):
            total_row = r
            break
    ok = total_row is not None
    results.append(('Total row (總計) found', ok,
                    f'Row {total_row}' if ok else 'Not found'))

    # 5 — Row 38 = POS
    ok = ws.cell(38, 1).value and 'POS' in str(ws.cell(38, 1).value)
    results.append(('Row 38 is POS', ok, f'A38={ws.cell(38,1).value}'))

    # 6 — Row 39 = Difference with formulas
    ok = ws.cell(39, 1).value and 'Diff' in str(ws.cell(39, 1).value)
    results.append(('Row 39 is Difference', ok, f'A39={ws.cell(39,1).value}'))
    has_formulas = any(str(ws.cell(39, c).value).startswith('=')
                       for c in range(2, 15) if ws.cell(39, c).value)
    results.append(('Row 39 has formulas', has_formulas, ''))

    # 7 — DBS dates in correct month
    dbs_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(4, c).value and 'DBS Receipt' == str(ws.cell(4, c).value):
            dbs_col = c
            break
    if dbs_col:
        bad = False
        for r in range(5, 38):
            v = ws.cell(r, dbs_col).value
            if v and str(v) != '總計':
                d = parse_date_str(v)
                if d and (d.year != year or d.month != month):
                    bad = True
                    break
        results.append(('DBS dates in correct month', not bad, ''))

    wb.close()
    passed = all(ok for _, ok, _ in results)
    return passed, results


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description='KPay Monthly Reconciliation')
    ap.add_argument('--pos',      required=True)
    ap.add_argument('--kpay',     required=True)
    ap.add_argument('--dbs',      required=True)
    ap.add_argument('--month',    required=True, help='YYYYMM')
    ap.add_argument('--store',    required=True)
    ap.add_argument('--merchant', required=True)
    ap.add_argument('--output',   required=True)
    args = ap.parse_args()

    year  = int(args.month[:4])
    month = int(args.month[4:6])

    print(f'\n{"="*60}')
    print(f'KPay Reconciliation — {args.store} — {args.month}')
    print(f'{"="*60}')

    print('\n[1/4] Reading KPay transactions…')
    kpay_data, settle_dates, methods_found = read_kpay(args.kpay)
    print(f'  → {len(settle_dates)} settle dates, methods: {sorted(methods_found)}')

    methods, rare = determine_methods(methods_found)
    print(f'  → Main grid ({len(methods)}): {methods}')
    if rare:
        print(f'  → Rare methods ({len(rare)}): {rare}')

    print('\n[2/4] Reading POS sales…')
    pos_daily, _ = read_pos(args.pos, args.store)
    print(f'  → {len(pos_daily)} trading days')

    print('\n[3/4] Reading DBS bank statement…')
    dbs_by_date, _ = read_dbs(args.dbs, args.merchant, year, month)
    print(f'  → {len(dbs_by_date)} credit dates, total HKD {sum(dbs_by_date.values()):,.2f}')

    print('\n[4/4] Building reconciliation report…')
    wb = Workbook()
    wb.remove(wb.active)
    layout, rows = build_sheet(wb, pos_daily, kpay_data, settle_dates,
                               dbs_by_date, args.store, year, month,
                               methods, rare)
    wb.save(args.output)
    print(f'  → Saved: {args.output}')
    print(f'  → {layout["num_methods"]} methods, total at row {rows["ROW_TOTAL"]}')

    print('\nValidation:')
    passed, val_results = validate_output(args.output, year, month)
    for name, ok, detail in val_results:
        sym = '✓' if ok else '✗'
        line = f'  {sym} {name}'
        if detail:
            line += f'  ({detail})'
        print(line)

    print(f'\n{"✓ All checks passed" if passed else "✗ Some checks failed"}')
    if not passed:
        sys.exit(1)


if __name__ == '__main__':
    main()

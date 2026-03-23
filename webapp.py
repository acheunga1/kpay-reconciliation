"""
KPay Reconciliation Web App — Flask backend
=============================================
Standalone web app for Wai Shing accounting staff.
Upload POS, KPay, and DBS files → auto-detect stores → batch reconcile → download zip.

Run with:  python webapp.py
Then open: http://localhost:5000
"""

import io
import os
import re
import sys
import tempfile
import zipfile
from datetime import datetime

from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook

# ── Import reconciliation engine ──────────────────────────────────────────────
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "tools"))
from batch_reconcile import SHOP_MID, MONTH_NAMES
from reconcile_kpay import (
    build_sheet,
    determine_methods,
    read_dbs,
    read_kpay,
    read_pos,
    validate_output,
)

app = Flask(__name__, template_folder="templates", static_folder="static")
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024  # 100 MB max upload


# ── File classification ───────────────────────────────────────────────────────

def classify_file(filename):
    """
    Classify an uploaded file as POS, KPay, or DBS based on filename patterns.
    Returns (file_type, shop_code_or_None).
    """
    name = filename.strip()

    # DBS bank statement: DBS_Sunstage_*.xls
    if name.upper().startswith("DBS_") or "SUNSTAGE" in name.upper():
        return "dbs", None

    # POS sales: POS Sales_YYYYMM.xlsx or POS_Sales_YYYYMM.xlsx
    if "POS" in name.upper() and "SALES" in name.upper():
        return "pos", None

    # KPay transaction report: #NN_SHOP_TERMINALID_YYYYMM.xlsx
    m = re.match(r"#\d+_([A-Za-z]+)_", name)
    if m:
        shop = m.group(1).upper()
        # Map known shop codes (e.g., "KYR" from "#02_KYR_...")
        if shop in SHOP_MID:
            return "kpay", shop
        # Try prefix match against known codes (longest first to avoid MI matching MIKI)
        for code in sorted(SHOP_MID.keys(), key=len, reverse=True):
            if shop.startswith(code):
                return "kpay", code
        return "kpay", shop

    return "unknown", None


def detect_shops_in_pos(pos_path):
    """Scan POS file to find which store codes are present."""
    import openpyxl as xl
    wb = xl.load_workbook(pos_path, read_only=True)
    for sheet_name in ["Sheet (2)", "Sheet"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            break
    else:
        ws = wb.active

    stores = set()
    for row in ws.iter_rows(min_row=3, max_col=1, values_only=True):
        v = str(row[0]).strip() if row[0] else ""
        if v and v != "總計" and "Total" not in v:
            stores.add(v)
    wb.close()
    return stores


def match_pos_store(pos_stores, shop_code):
    """Match a shop code to a POS store name (handles MI→MIKI, etc.)."""
    if shop_code in pos_stores:
        return shop_code
    candidates = [s for s in pos_stores if s.startswith(shop_code)]
    if len(candidates) == 1:
        return candidates[0]
    if candidates:
        return min(candidates, key=len)
    return None


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html",
                           shops=sorted(SHOP_MID.keys()),
                           months=MONTH_NAMES,
                           current_year=datetime.now().year)


@app.route("/api/reconcile", methods=["POST"])
def api_reconcile():
    """
    Accept uploaded files + month/year, run reconciliation, return zip.

    Expected form fields:
        month: int (1-12)
        year:  int (e.g. 2026)
        files: multiple file uploads
    """
    try:
        month = int(request.form.get("month", 0))
        year = int(request.form.get("year", 0))
        if not (1 <= month <= 12) or not (2024 <= year <= 2030):
            return jsonify({"error": "Invalid month or year"}), 400

        files = request.files.getlist("files")
        if not files:
            return jsonify({"error": "No files uploaded"}), 400

        month_str = f"{year}{month:02d}"
        mon_name = MONTH_NAMES[month - 1]

        # ── Save uploads to temp dir and classify ─────────────────────────
        with tempfile.TemporaryDirectory() as tmpdir:
            pos_path = None
            dbs_path = None
            kpay_files = {}   # {shop_code: path}
            unknown_files = []

            for f in files:
                if not f.filename:
                    continue
                ftype, shop = classify_file(f.filename)
                save_path = os.path.join(tmpdir, f.filename)
                f.save(save_path)

                if ftype == "pos":
                    pos_path = save_path
                elif ftype == "dbs":
                    dbs_path = save_path
                elif ftype == "kpay":
                    if shop:
                        kpay_files[shop] = save_path
                    else:
                        unknown_files.append(f.filename)
                else:
                    unknown_files.append(f.filename)

            # ── Validate we have the minimum files ────────────────────────
            errors = []
            if not pos_path:
                errors.append("Missing POS Sales file")
            if not dbs_path:
                errors.append("Missing DBS bank statement file")
            if not kpay_files:
                errors.append("No KPay transaction files detected")
            if errors:
                return jsonify({"error": "; ".join(errors),
                                "unknown_files": unknown_files}), 400

            # ── Detect stores available in POS ────────────────────────────
            pos_stores = detect_shops_in_pos(pos_path) if pos_path else set()

            # ── Process each store ────────────────────────────────────────
            results = []
            output_dir = os.path.join(tmpdir, "output")
            os.makedirs(output_dir, exist_ok=True)

            for shop, kpay_path in sorted(kpay_files.items()):
                shop_result = {"shop": shop, "status": "pending", "checks": []}

                merchant = SHOP_MID.get(shop)
                if not merchant:
                    shop_result["status"] = "skipped"
                    shop_result["error"] = f"Unknown shop code: {shop}"
                    results.append(shop_result)
                    continue

                # Check if POS has data for this store
                pos_store = match_pos_store(pos_stores, shop)
                if not pos_store:
                    shop_result["status"] = "warning"
                    shop_result["error"] = (
                        f"Shop {shop} not found in POS file. "
                        f"Available: {sorted(pos_stores)}"
                    )

                try:
                    # 1. Read KPay
                    kpay_data, settle_dates, methods_found = read_kpay(kpay_path)
                    methods, rare = determine_methods(methods_found)

                    # 2. Read POS
                    pos_daily, _ = read_pos(pos_path, shop)

                    # 3. Read DBS
                    dbs_by_date, _ = read_dbs(dbs_path, merchant, year, month)

                    # 4. Build workbook
                    wb = Workbook()
                    wb.remove(wb.active)
                    build_sheet(wb, pos_daily, kpay_data, settle_dates,
                                dbs_by_date, shop, year, month, methods, rare)

                    out_name = f"{shop}_KPay_{mon_name}{year}_Reconciliation.xlsx"
                    out_path = os.path.join(output_dir, out_name)
                    wb.save(out_path)

                    # 5. Validate
                    passed, checks = validate_output(out_path, year, month)
                    shop_result["status"] = "pass" if passed else "fail"
                    shop_result["checks"] = [
                        {"name": n, "ok": ok, "detail": d}
                        for n, ok, d in checks
                    ]
                    shop_result["file"] = out_name
                    shop_result["summary"] = {
                        "settle_dates": len(settle_dates),
                        "methods": len(methods_found),
                        "rare_methods": rare,
                        "dbs_credits": len(dbs_by_date),
                        "dbs_total": round(sum(dbs_by_date.values()), 2),
                        "pos_days": len(pos_daily),
                    }

                except Exception as e:
                    shop_result["status"] = "error"
                    shop_result["error"] = str(e)

                results.append(shop_result)

            # ── Package outputs as zip ────────────────────────────────────
            output_files = [f for f in os.listdir(output_dir)
                            if f.endswith(".xlsx")]

            if not output_files:
                return jsonify({"error": "No reports generated",
                                "results": results}), 500

            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                for fname in output_files:
                    fpath = os.path.join(output_dir, fname)
                    zf.write(fpath, fname)
            zip_buffer.seek(0)

            # If only one file, return the Excel directly
            if len(output_files) == 1:
                single_path = os.path.join(output_dir, output_files[0])
                with open(single_path, "rb") as f:
                    file_bytes = f.read()

                return jsonify({
                    "results": results,
                    "download_ready": True,
                    "filename": output_files[0],
                    "file_count": 1,
                })

            # Store zip in a temp file for download
            zip_name = f"KPay_Reconciliation_{mon_name}{year}.zip"
            zip_path = os.path.join(tmpdir, zip_name)
            with open(zip_path, "wb") as f:
                f.write(zip_buffer.getvalue())

            # Return results first, then client fetches the download
            # We need to save the zip outside tmpdir so it persists
            persist_dir = os.path.join(os.path.dirname(__file__), ".tmp", "downloads")
            os.makedirs(persist_dir, exist_ok=True)
            persist_path = os.path.join(persist_dir, zip_name)
            with open(persist_path, "wb") as f:
                f.write(zip_buffer.getvalue())

            # Also save individual files for single download
            for fname in output_files:
                src = os.path.join(output_dir, fname)
                dst = os.path.join(persist_dir, fname)
                with open(src, "rb") as sf:
                    with open(dst, "wb") as df:
                        df.write(sf.read())

            return jsonify({
                "results": results,
                "download_ready": True,
                "filename": zip_name,
                "file_count": len(output_files),
                "individual_files": output_files,
            })

    except Exception as e:
        return jsonify({"error": f"Server error: {str(e)}"}), 500


@app.route("/api/download/<filename>")
def api_download(filename):
    """Download a generated report file."""
    # Sanitize filename
    filename = os.path.basename(filename)
    persist_dir = os.path.join(os.path.dirname(__file__), ".tmp", "downloads")
    filepath = os.path.join(persist_dir, filename)

    if not os.path.exists(filepath):
        return jsonify({"error": "File not found"}), 404

    if filename.endswith(".zip"):
        mimetype = "application/zip"
    else:
        mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return send_file(filepath, as_attachment=True,
                     download_name=filename, mimetype=mimetype)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5050, debug=True)
